
import requests
from bs4 import BeautifulSoup
import openpyxl as xl

wb = xl.load_workbook('C://Users//John Berry//Desktop//VA_Links.xlsx')
ws = wb.get_sheet_by_name("Sheet1")
FACILITY_NAME_COL = 1
LINK_COL = 2
ROW_COUNTER = 2


target_url_base = "http://www.dss.virginia.gov/facility/search/alf.cgi?rm=Search;search_modifiers_assisted_living=ASST;Start"

facility_urls = []
to_visit = []
visited_urls = []

visit_counter = 0


for i in range(1,484,25):
    to_visit.append(str(i))

def call_url():
    FACILITY_NAME_COL = 1
    LINK_COL = 2
    ROW_COUNTER = 2
    VISIT_COUNTER = 0

    while VISIT_COUNTER != len(to_visit):
        scraping_url = target_url_base + to_visit[VISIT_COUNTER] ## PUT IN DYNAMIC COUNTER
        #visit_counter +=1
        r = requests.get(scraping_url)
        scraped_html = r.text
        soup = BeautifulSoup(scraped_html, "html.parser")
        linked_tables = soup.findAll('td', valign="TOP")

        for i in linked_tables:
            link = i.find('a', href=True)
            link = link['href']
            text = i.text

            ws.cell(row = ROW_COUNTER, column = FACILITY_NAME_COL).value = text
            ws.cell(row = ROW_COUNTER, column = LINK_COL).value = link
            
            ROW_COUNTER +=1

        VISIT_COUNTER+=1

    wb.save('C://Users//John Berry//Desktop//VA_Links1.xlsx')
    print("All Done!")

            
        
       





    
    

    

    
