import requests
import pprint
from bs4 import BeautifulSoup
import openpyxl as xl
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By

clean_detail_data = []
LINK_COL = 2
ROW_COUNT = 2
WRITE_COUNTER = 3

wb = xl.load_workbook('VA_Links.xlsx')
ws = wb.get_sheet_by_name('Sheet1')

target_wb = wb = xl.load_workbook('Virgina_DB.xlsx')
target_ws = target_wb.get_sheet_by_name('DATA')

TARGET_BASE_URL = "http://www.dss.virginia.gov"

driver = webdriver.Chrome('C:\\Users\\John Berry\\Desktop\\Virginia AL Scraper\\chromedriver.exe')

def selenium_scraper(url):
    driver.get(url)
    rows = driver.find_elements_by_tag_name('tr')

    for row in rows:
        row = row.text.strip()
        row = row.replace("\n"," ")
        
        row = row.replace("Facility Type: ","")
        row = row.replace("License Type: ","")
        row = row.replace("Expiration Date: ","")
        row = row.replace("Qualification: ","")
        row = row.replace("Administrator: ","")
        row = row.replace("Business Hours: ","")
        row = row.replace("Capacity: ","")
        row = row.replace("Inspector: ","")
        
        clean_detail_data.append(row)
    try:
       excelWriter(clean_detail_data)
    except:
        target_wb.save("VA_run1.xlsx")
        print("ENCOUNTERED ERROR WHILE WRITING")
        driver.quit()
        
def excelWriter(data):
    
    global WRITE_COUNTER
    
    for i in range(0,12):
        target_ws.cell(row=WRITE_COUNTER, column=i+1).value = data[i]

    WRITE_COUNTER+=1
    

while ROW_COUNT != 488:
    clean_detail_data = []

    fac_link = ws.cell(row=ROW_COUNT, column=LINK_COL).value
    url_to_scrape = TARGET_BASE_URL + fac_link

    selenium_scraper(url_to_scrape)
    target_wb.save("VA_run1.xlsx")
    
    ROW_COUNT+=1

driver.quit()

